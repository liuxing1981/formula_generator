import json

from flask import Flask, jsonify, request, Response, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, colors
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from werkzeug.utils import redirect

from calculate import calculate

app = Flask(__name__)
with open ('models.json',encoding='utf-8') as f:
    models = json.load(f)

@app.route('/')
def index():
    return redirect(url_for('static', filename='main.html'))

@app.route('/model')
def model():
    return jsonify(models)

@app.route('/generateFormula',methods=['POST'])
def generateFormula():
    result = set()
    data = request.json['data']
    if data[0]['model'] == '@@':
        custom = []
        number = int(data[0]['number'])
        add = number // 5
        mul = number - add
        cal = calculate('2,9 * 2,9 = 100', number)
        cal.do_task()
        for index, r in enumerate(cal.result):
            if (index + 1) % 5 != 0:
                custom.append(r)
            else:
                custom.append('@@')
        cal = calculate('2,9 * 2,9 +- 50= 100', add)
        cal.do_task()
        for index in range(0, len(custom)):
            if custom[index] == '@@':
                custom[index] = cal.result.pop()
        return jsonify({'data': ('@').join(custom)})
    else:
        for element in data:
            cal = calculate(element['model'], int(element['number']))
            cal.do_task()
            result.update(cal.result)
    return jsonify({'data': ('@').join(result)})

@app.route('/downloadFormula',methods=['POST'])
def downloadFormula():
    data = eval(request.json['data'])
    data_length = len(data)
    font = Font('宋体', size=14, bold=False, italic=False, strike=False, color=colors.BLACK)
    align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    wb = Workbook()
    page = 1
    for k in range(page):
        if k == 0:
            ws = wb.active
            ws.title = 'page1'
        else:
            ws = wb.create_sheet(index=k, title='page' + str(k + 1))
        nameCell = ws.cell(row=1, column=4, value='姓名：')
        classCell = ws.cell(row=1, column=5, value='班级：')
        nameCell.font = font
        classCell.font = font
        splitLine = data_length // 2 + 3
        for i in range(3, data_length + 3):
            for j in range(1, len(data[i - 3]) + 1):
                # insert a empty line for split to half
                if i == splitLine and data_length > 5:
                    ws.cell(row=i, column=j)
                    continue
                myCell = ws.cell(row=i, column=j, value=data[i - 3][j - 1])
                myCell.font = font
                myCell.alignment = align
                # print('write to %s,%s %s'%(i,j,content))
        for i in range(1, ws.max_row + 1):
            ws.row_dimensions[i].height = 30
        for i in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16
    return Response(
        save_virtual_workbook(wb),
        headers={
            'Content-Disposition': 'attachment; filename=sheet.xlsx',
            'Content-type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })


if __name__ == '__main__':
    app.run(host='127.0.0.1', port=80, debug=True)